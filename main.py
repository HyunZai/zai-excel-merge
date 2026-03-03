import sys
import os
import shutil
from datetime import datetime
from PySide6.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, 
                               QLabel, QWidget, QFileDialog, QListWidget, QTextEdit, QHBoxLayout, QLineEdit)
from PySide6.QtCore import Qt, QThread, Signal
import openpyxl

# --- 1-1. 병합 처리 워커 스레드 ---
class MergeWorker(QThread):
    log_signal = Signal(str)
    finished_signal = Signal(bool)

    def __init__(self, file_paths):
        super().__init__()
        self.file_paths = file_paths

    def run(self):
        try:
            total_files = len(self.file_paths)
            if total_files == 0:
                self.finished_signal.emit(False)
                return

            self.log_signal.emit("🔍 파일들을 스캔하여 가장 컬럼이 많은 기준 파일을 찾는 중...")
            
            max_cols = 0
            base_file_path = self.file_paths[0]
            base_file_index = 0
            
            for idx, file_path in enumerate(self.file_paths):
                try:
                    wb_temp = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws_temp = wb_temp.active
                    current_cols = ws_temp.max_column
                    wb_temp.close()
                    
                    if current_cols > max_cols:
                        max_cols = current_cols
                        base_file_path = file_path
                        base_file_index = idx
                except Exception as e:
                    self.log_signal.emit(f"⚠️ {os.path.basename(file_path)} 파일 스캔 실패: {e}")

            base_filename = os.path.basename(base_file_path)
            base_dir = os.path.dirname(self.file_paths[0])
            
            parent_folder_name = os.path.basename(base_dir)
            if not parent_folder_name: 
                parent_folder_name = "merged"
                
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"{parent_folder_name}_{current_time}.xlsx"
            result_path = os.path.join(base_dir, new_filename)

            shutil.copy2(base_file_path, result_path)
            
            base_wb = openpyxl.load_workbook(result_path)
            base_ws = base_wb.active
            base_rows = base_ws.max_row
            
            self.log_signal.emit(f"✅ 기준 파일 설정 완료: 1번째 파일({base_filename}) - 데이터 행: {max(0, base_rows - 1)}, 열: {max_cols}\n" + "-"*40)

            remaining_files = [f for i, f in enumerate(self.file_paths) if i != base_file_index]
            
            processed_count = 1
            
            for current_file in remaining_files:
                processed_count += 1
                current_filename = os.path.basename(current_file)
                file_idx_str = f"[{processed_count:02d}/{total_files:02d}]"

                try:
                    wb = openpyxl.load_workbook(current_file, data_only=True)
                    ws = wb.active
                    
                    c_rows = ws.max_row
                    c_cols = ws.max_column

                    self.log_signal.emit(f"{file_idx_str} 작업 중... ({current_filename}) - 데이터 행: {max(0, c_rows - 1)}, 열: {c_cols}")

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            base_ws.append(row)

                    wb.close()
                    self.log_signal.emit(f"{file_idx_str} 작업 완료")

                except Exception as e:
                    self.log_signal.emit(f"{file_idx_str} 파일 읽기에 실패했습니다. [{current_filename}] 파일을 확인해주세요. (에러: {e})")

            base_wb.save(result_path)
            self.log_signal.emit(f"\n🎉 모든 병합 작업이 완료되었습니다!\n저장 파일명: {new_filename}")
            self.finished_signal.emit(True)

        except Exception as e:
            self.log_signal.emit(f"❌ 치명적인 오류 발생: {e}")
            self.finished_signal.emit(False)

# --- 1-2. 수험번호 검색 워커 스레드 (새로 추가됨) ---
class SearchWorker(QThread):
    log_signal = Signal(str)
    finished_signal = Signal(bool)

    def __init__(self, file_paths, keyword):
        super().__init__()
        self.file_paths = file_paths
        self.keyword = str(keyword).strip() # 공백 제거 후 문자열 처리

    def run(self):
        try:
            self.log_signal.emit(f"\n🔎 '{self.keyword}' 수험번호 검색을 시작합니다...")
            found_files = []

            for file_path in self.file_paths:
                filename = os.path.basename(file_path)
                try:
                    # 데이터만 빠르게 읽기 위해 read_only=True 사용
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    found_in_file = False
                    
                    for row in ws.iter_rows(values_only=True):
                        # 행 안의 셀 데이터들을 문자열로 변환하여 검색어와 일치하는지 확인
                        if any(self.keyword == str(cell).strip() for cell in row if cell is not None):
                            found_in_file = True
                            break # 해당 파일에서 찾았으면 다음 파일로 넘어감
                            
                    wb.close()
                    
                    if found_in_file:
                        found_files.append(filename)
                        self.log_signal.emit(f"🎯 [발견] {filename} 파일에 해당 수험번호가 있습니다.")
                except Exception as e:
                    self.log_signal.emit(f"⚠️ {filename} 읽기 실패: {e}")

            if not found_files:
                self.log_signal.emit(f"❌ 결과: 선택된 파일 중 '{self.keyword}' 수험번호를 찾을 수 없습니다.")
            else:
                self.log_signal.emit(f"✅ 검색 완료: 총 {len(found_files)}개의 파일에서 발견되었습니다.")
            
            self.finished_signal.emit(True)

        except Exception as e:
            self.log_signal.emit(f"❌ 검색 중 오류 발생: {e}")
            self.finished_signal.emit(False)


# --- 2. 화면(UI)을 구성하는 메인 윈도우 ---
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 병합 및 검색 프로그램")
        self.setMinimumSize(700, 550) # 입력란이 추가되어 가로 길이를 조금 늘렸습니다.
        self.selected_files = []

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # 상단: 파일 선택
        top_layout = QHBoxLayout()
        self.btn_select = QPushButton("엑셀 파일 선택하기", self)
        self.btn_select.setMinimumHeight(40)
        self.btn_select.clicked.connect(self.select_files)
        top_layout.addWidget(self.btn_select)

        self.label_count = QLabel("선택된 파일: 0개", self)
        self.label_count.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        top_layout.addWidget(self.label_count)
        layout.addLayout(top_layout)

        # 중단: 리스트
        self.list_widget = QListWidget(self)
        layout.addWidget(self.list_widget)

        # 중단 액션바: 병합 버튼 + 검색 입력란 + 검색 버튼
        action_layout = QHBoxLayout()
        
        self.btn_merge = QPushButton("병합 시작", self)
        self.btn_merge.setMinimumHeight(40)
        self.btn_merge.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        self.btn_merge.clicked.connect(self.start_merge)
        self.btn_merge.setEnabled(False) 
        action_layout.addWidget(self.btn_merge, stretch=2) # 너비 비율 조정

        self.input_search = QLineEdit(self)
        self.input_search.setMinimumHeight(40)
        self.input_search.setPlaceholderText("검색 키워드 입력")
        self.input_search.setEnabled(False)
        # 엔터키를 눌러도 검색되도록 연결
        self.input_search.returnPressed.connect(self.start_search)
        action_layout.addWidget(self.input_search, stretch=2)

        self.btn_search = QPushButton("검색", self)
        self.btn_search.setMinimumHeight(40)
        self.btn_search.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        self.btn_search.clicked.connect(self.start_search)
        self.btn_search.setEnabled(False)
        action_layout.addWidget(self.btn_search, stretch=1)
        
        layout.addLayout(action_layout)

        # 하단: 로그 콘솔
        self.log_console = QTextEdit(self)
        self.log_console.setReadOnly(True)
        self.log_console.setStyleSheet("background-color: #1E1E1E; color: #00FF00; font-family: monospace; font-size: 13px;")
        layout.addWidget(self.log_console)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls)")
        
        if files:
            # --- 🌟 신규 작업 전 완벽 초기화 로직 ---
            self.selected_files.clear() # 혹시 모를 내부 리스트 찌꺼기 비우기
            self.list_widget.clear()    # 화면의 파일 리스트 지우기
            self.log_console.clear()    # 로그창 텍스트 지우기
            self.input_search.clear()   # 검색어 입력란 비우기
            # ------------------------------------
            
            self.selected_files = files
            self.label_count.setText(f"선택된 파일: {len(self.selected_files)}개")
            
            for f in self.selected_files:
                self.list_widget.addItem(os.path.basename(f))
            
            # 파일 선택 후 모든 기능 버튼 활성화
            self.btn_merge.setEnabled(True)
            self.input_search.setEnabled(True)
            self.btn_search.setEnabled(True)
            
            self.log_console.append("✨ 이전 작업 내역이 초기화되었습니다.")
            self.log_console.append(f"✅ {len(self.selected_files)}개의 파일이 새로 선택되었습니다. 병합을 시작하거나 수험번호를 검색해보세요.")

    def start_merge(self):
        self.btn_select.setEnabled(False)
        self.btn_merge.setEnabled(False)
        self.btn_search.setEnabled(False)
        self.input_search.setEnabled(False)
        self.log_console.clear()
        
        self.merge_worker = MergeWorker(self.selected_files)
        self.merge_worker.log_signal.connect(self.update_log)
        self.merge_worker.finished_signal.connect(self.action_finished)
        self.merge_worker.start()

    def start_search(self):
        keyword = self.input_search.text()
        if not keyword:
            self.update_log("⚠️ 검색할 수험번호를 입력해주세요.")
            self.input_search.setFocus()
            return

        self.btn_select.setEnabled(False)
        self.btn_merge.setEnabled(False)
        self.btn_search.setEnabled(False)
        self.input_search.setEnabled(False)
        
        self.search_worker = SearchWorker(self.selected_files, keyword)
        self.search_worker.log_signal.connect(self.update_log)
        self.search_worker.finished_signal.connect(self.action_finished)
        self.search_worker.start()

    def update_log(self, message):
        self.log_console.append(message)
        self.log_console.ensureCursorVisible()

    def action_finished(self, success):
        self.btn_select.setEnabled(True)
        self.btn_merge.setEnabled(True)
        self.btn_search.setEnabled(True)
        self.input_search.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())